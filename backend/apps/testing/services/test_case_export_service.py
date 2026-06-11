"""测试用例 Excel 导出。"""

from __future__ import annotations

import io
import json
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.testing.models import TestCase

BASE_COLUMNS: list[tuple[str, str, int]] = [
    ("index", "序号", 8),
    ("case_no", "用例标号", 14),
    ("title", "用例标题", 28),
    ("module", "模块", 14),
    ("project_name", "所属项目", 16),
    ("test_point_name", "测试点", 16),
    ("priority", "优先级", 10),
    ("precondition", "前置条件", 24),
    ("steps", "测试步骤", 36),
    ("expected", "预期结果", 28),
    ("postcondition", "后置条件", 20),
    ("executor", "执行人", 12),
    ("passed_label", "执行状态", 10),
    ("actual", "实际结果", 24),
    ("source_type_label", "用例来源", 12),
    ("created_at", "创建时间", 20),
]

SOURCE_LABELS = {
    "manual": "手动用例",
    "ai": "AI用例",
    "test_point": "测试点生成",
    "template": "模板生成",
}


def _passed_label(passed: bool | None) -> str:
    if passed is True:
        return "已通过"
    if passed is False:
        return "已失败"
    return "未执行"


def _collect_extra_keys(cases: list[TestCase]) -> list[str]:
    keys: list[str] = []
    seen: set[str] = set()
    for case in cases:
        for key in (case.extra_data or {}).keys():
            if key not in seen:
                seen.add(key)
                keys.append(key)
    return keys


def _row_values(case: TestCase, index: int, extra_keys: list[str]) -> list[Any]:
    row = [
        index,
        case.case_no or "",
        case.title or "",
        case.module or "",
        case.project.name if case.project_id else "",
        case.test_point.name if case.test_point_id else "",
        case.priority or "",
        case.precondition or "",
        case.steps or "",
        case.expected or "",
        case.postcondition or "",
        case.executor or "",
        _passed_label(case.passed),
        case.actual or "",
        SOURCE_LABELS.get(case.source_type, case.source_type or ""),
        case.created_at.strftime("%Y-%m-%d %H:%M:%S") if case.created_at else "",
    ]
    extra = case.extra_data or {}
    for key in extra_keys:
        val = extra.get(key, "")
        if isinstance(val, (dict, list)):
            val = json.dumps(val, ensure_ascii=False)
        row.append(val if val is not None else "")
    return row


def build_test_cases_excel(cases: list[TestCase]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    extra_keys = _collect_extra_keys(cases)
    headers = [col[1] for col in BASE_COLUMNS] + extra_keys
    widths = [col[2] for col in BASE_COLUMNS] + [16] * len(extra_keys)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)

    ws.append(headers)
    for col_idx, width in enumerate(widths, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for index, case in enumerate(cases, start=1):
        ws.append(_row_values(case, index, extra_keys))

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="top")
        for col_idx in range(2, len(headers) + 2):
            ws.cell(row=row_idx, column=col_idx).alignment = body_alignment

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_export_filename(project_name: str = "") -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = "".join(c for c in project_name if c.isalnum() or c in ("-", "_", " ")).strip()
    if safe_name:
        return f"测试用例_{safe_name}_{stamp}.xlsx"
    return f"测试用例_{stamp}.xlsx"
